#!/usr/bin/env python3
"""
Clean the *external* research sources (market size, Mintel survey, Reddit) from
the raw corpus into small committed CSVs under data/market/ and data/reddit/.

Like build_clean_datasets.py, the raw corpus is NOT committed — point this at an
unzipped copy and it regenerates the cleaned CSVs that build_dashboard_json.py
then folds into the dashboard aggregate.

  pip install openpyxl vaderSentiment
  RAW_DATA_DIR=/path/to/02_Data python data/scripts/build_external_datasets.py

Expects under RAW_DATA_DIR:
  Catalyst Datasets/consumer_..._worldwide_USD_en.xlsx   (Statista 'Revenue' sheet)
  Mintel 2026 Energy Report Data/...Full Databook.xlsx    ('Q6 topline', 'Q4 topline')
  Reddit data/r-Energy Drinks/energydrinks_{posts,comments}_20260609.csv
"""
import csv
import os
import re
import sys
import glob

import openpyxl

csv.field_size_limit(10**9)

RAW = os.environ.get("RAW_DATA_DIR", "/tmp/corpus/02_Data")
ROOT = os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
DATA = os.path.join(ROOT, "data")


def find(*parts):
    """Glob-match a file under RAW (sheet names/paths have odd unicode)."""
    hits = glob.glob(os.path.join(RAW, *parts))
    return hits[0] if hits else None


def write_csv(relpath, header, rows):
    path = os.path.join(DATA, relpath)
    os.makedirs(os.path.dirname(path), exist_ok=True)
    with open(path, "w", encoding="utf-8", newline="") as fh:
        w = csv.writer(fh)
        w.writerow(header)
        w.writerows(rows)
    print(f"  wrote {relpath} ({len(rows)} rows)")


# ---------------------------------------------------------------- market size (Catalyst / Statista)
def build_market():
    f = find("Catalyst Datasets", "*worldwide_USD*.xlsx")
    if not f:
        print("  ! Catalyst worldwide USD file not found", file=sys.stderr)
        return
    wb = openpyxl.load_workbook(f, read_only=True, data_only=True)
    ws = wb["Revenue"]
    grid = [list(r) for r in ws.iter_rows(values_only=True)]

    def row_after(label_sub):
        for i, r in enumerate(grid):
            if r and isinstance(r[0], str) and label_sub.lower() in r[0].lower():
                return i
        return None

    def years_in(row):
        out = []
        for c in row[1:]:
            if isinstance(c, int) and 1990 < c < 2100:
                out.append(c)
            elif isinstance(c, str) and re.fullmatch(r"(19|20)\d{2}", c.strip()):
                out.append(int(c.strip()))
        return out

    def nums_in(row):
        return [c for c in row[1:] if isinstance(c, (int, float)) and not isinstance(c, bool)]

    # AT-HOME revenue: a "years" row then a "Total" row
    i = row_after("REVENUE, AT HOME")
    size = list(zip(years_in(grid[i + 1]), nums_in(grid[i + 2])))

    # YoY change %
    j = row_after("REVENUE CHANGE")
    chg = dict(zip(years_in(grid[j + 1]), nums_in(grid[j + 2])))

    # out-of-home 2024
    k = row_after("OUT-OF-HOME")
    ooh = next((c for c in grid[k + 2][1:] if isinstance(c, (int, float))), None)

    rows = [(y, round(v, 2), chg.get(y, "")) for y, v in size]
    write_csv("market/market_size.csv", ["year", "revenue_busd", "yoy_change_pct"], rows)

    by_year = dict(size)
    y0, v0 = size[0]
    a2024 = by_year.get(2024, size[-1][1])
    yf, vf = size[-1]  # last forecast year
    cagr_hist = ((a2024 / v0) ** (1 / (2024 - y0)) - 1) * 100
    cagr_fwd = ((vf / a2024) ** (1 / (yf - 2024)) - 1) * 100 if yf > 2024 else None
    facts = [
        ("athome_2024_busd", round(a2024, 2)),
        ("ooh_2024_busd", round(ooh, 2) if ooh else ""),
        ("total_2024_busd", round(a2024 + (ooh or 0), 2)),
        ("forecast_year", yf),
        ("forecast_busd", round(vf, 2)),
        ("cagr_hist_pct", round(cagr_hist, 1)),
        ("cagr_hist_span", f"{y0}-2024"),
        ("cagr_fwd_pct", round(cagr_fwd, 1) if cagr_fwd else ""),
        ("cagr_fwd_span", f"2024-{yf}"),
        ("source", "Statista Market Insights (Catalyst), upd 03/2026"),
    ]
    write_csv("market/market_facts.csv", ["key", "value"], facts)


# ---------------------------------------------------------------- Mintel survey (concept interest + motivations)
def build_mintel():
    f = find("Mintel*", "*Full Databook*.xlsx")
    if not f:
        print("  ! Mintel databook not found", file=sys.stderr)
        return
    wb = openpyxl.load_workbook(f, read_only=True, data_only=True)

    # Q6 — concept interest: rows of (concept, share 0..1)
    ws = wb["Q6 topline"]
    concepts = []
    for r in ws.iter_rows(values_only=True):
        label, val = (r[0] if r else None), (r[1] if len(r) > 1 else None)
        if isinstance(label, str) and isinstance(val, (int, float)) and 0 < val <= 1:
            if label.strip().lower() in ("sample", "none of these"):
                continue
            concepts.append((label.strip(), round(val * 100, 1)))
    concepts.sort(key=lambda x: -x[1])
    write_csv("market/concept_interest.csv", ["concept", "interest_pct"], concepts)

    # Q4 — motivating factors: top-2-box (Extremely + Very) per factor column
    ws = wb["Q4 topline"]
    grid = [list(r) for r in ws.iter_rows(values_only=True)]
    hdr_i = next((i for i, r in enumerate(grid)
                  if r and isinstance(r[0], str) and r[0].strip().lower() == "sample"), None)
    if hdr_i is not None:
        factors = grid[hdr_i - 1]
        ext = next((r for r in grid if r and str(r[0]).strip().lower().startswith("extremely")), [])
        very = next((r for r in grid if r and str(r[0]).strip().lower().startswith("very")), [])
        rows = []
        for c in range(1, len(factors)):
            name = factors[c]
            e = ext[c] if c < len(ext) and isinstance(ext[c], (int, float)) else 0
            v = very[c] if c < len(very) and isinstance(very[c], (int, float)) else 0
            if isinstance(name, str) and (e or v):
                rows.append((name.strip(), round((e + v) * 100, 1)))
        rows.sort(key=lambda x: -x[1])
        write_csv("market/motivations.csv", ["factor", "top2box_pct"], rows)


# ---------------------------------------------------------------- Reddit (brand mentions + sentiment snapshot)
def build_reddit():
    base = find("Reddit data", "r-Energy Drinks")
    if not base:
        print("  ! Reddit folder not found", file=sys.stderr)
        return
    posts_f = os.path.join(base, "energydrinks_posts_20260609.csv")
    comments_f = os.path.join(base, "energydrinks_comments_20260609.csv")

    # brand list + aliases (from the cross-platform table)
    brands = []
    with open(os.path.join(DATA, "combined/brand_summary.csv"), encoding="utf-8") as fh:
        for r in csv.DictReader(fh):
            brands.append(r["brand"])
    ALIAS = {
        "Red Bull": r"red\s*bull", "G Fuel": r"g[\s-]?fuel", "Alani Nu": r"alani(\s*nu)?",
        "5-hour Energy": r"5[\s-]?hour", "Liquid I.V.": r"liquid\s*i\.?v\.?",
        "Liquid Death": r"liquid\s*death", "Bloom Nutrition": r"bloom",
        "Prime": r"\bprime\b", "Ghost": r"\bghost\b", "Reign": r"\breign\b",
        "Bang": r"\bbang\b", "C4": r"\bc4\b", "Zoa": r"\bzoa\b", "NOS": r"\bnos\b",
    }
    pats = [(b, re.compile(ALIAS.get(b, r"\b" + re.escape(b) + r"\b"), re.I)) for b in brands]

    try:
        from vaderSentiment.vaderSentiment import SentimentIntensityAnalyzer
        sia = SentimentIntensityAnalyzer()

        def sent(t):
            c = sia.polarity_scores(t)["compound"]
            return "pos" if c >= 0.05 else "neg" if c <= -0.05 else "neu"
        method = "VADER"
    except Exception:
        def sent(t):
            return "neu"
        method = "none"

    from collections import Counter, defaultdict
    mentions = Counter()
    sentic = defaultdict(Counter)
    dates = []

    def scan(text, ts):
        if not text:
            return
        hit = False
        for b, rx in pats:
            if rx.search(text):
                mentions[b] += 1
                sentic[b][sent(text)] += 1
                hit = True
        if hit and ts:
            try:
                dates.append(int(float(ts)))
            except ValueError:
                pass

    n_posts = n_comments = 0
    with open(posts_f, encoding="utf-8", errors="replace") as fh:
        for r in csv.DictReader(fh):
            n_posts += 1
            scan((r.get("title") or "") + " " + (r.get("selftext") or ""), r.get("created_utc"))
    with open(comments_f, encoding="utf-8", errors="replace") as fh:
        for r in csv.DictReader(fh):
            n_comments += 1
            scan(r.get("body"), r.get("created_utc"))

    rows = sorted(
        [(b, mentions[b], sentic[b]["pos"], sentic[b]["neg"], sentic[b]["neu"])
         for b in mentions if mentions[b] > 0],
        key=lambda x: -x[1],
    )
    write_csv("reddit/brand_pulse.csv", ["brand", "mentions", "pos", "neg", "neu"], rows)

    from datetime import datetime
    span = (f"{datetime.utcfromtimestamp(min(dates)).date()}",
            f"{datetime.utcfromtimestamp(max(dates)).date()}") if dates else ("", "")
    write_csv("reddit/meta.csv", ["key", "value"], [
        ("posts", n_posts), ("comments", n_comments),
        ("date_start", span[0]), ("date_end", span[1]),
        ("sentiment_method", method), ("subreddit", "r/EnergyDrinks"),
    ])


if __name__ == "__main__":
    print(f"Reading raw corpus from {RAW} ...")
    build_market()
    build_mintel()
    build_reddit()
    print("Done. Now run: python data/scripts/build_dashboard_json.py")
